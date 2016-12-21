from __future__ import print_function
import operator
import itertools
import collections
import math
import re
import openpyxl
import json

#=============================================================================================
# Reading query ids from file
# Query_ids -> 12, 13, 19

query_ids = []
with open("query_ids.txt") as f:
        lines = f.read().splitlines()

count = 0
for line in lines:
    string_list = line.split()
    query_ids.append(string_list[0])
    count += 1

#=============================================================================================
# reading query relevance level from given document.

query_relevance_levels = dict((query_id, {}) for query_id in query_ids)

with open("relevance_levels.txt") as f:
        relevance_lines = f.read().splitlines()

for query_id in query_ids:
    temp_dict = {}
    for line in relevance_lines:
        string_list = line.split()
        if query_id == string_list[0]:         
            doc_id = string_list[2]
            relevance_level = string_list[3]
            temp_dict[doc_id] = relevance_level
    query_relevance_levels[query_id] = temp_dict       
            
#=============================================================================================
# computing no. of relevant docs for each query.
relevant_docs_count = {}

for query_id in query_relevance_levels.keys():
    relevant_docs_count[query_id] = len(query_relevance_levels.get(query_id))

#=============================================================================================    
# computing recall and precision for each document and each query.

document_recall_dict = dict((query_id, {}) for query_id in query_ids)
document_precision_dict = dict((query_id, {}) for query_id in query_ids)

for query_id in query_ids:    
    filename = 'Query_' + query_id + '_results.xlsx'
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name('Sheet1')

    temp_recall_dict = {}
    temp_precision_dict = {}    
    relevant_docs = query_relevance_levels.get(query_id).keys()

    summation = 0
    for r in range(2, sheet.max_row + 1):
        rank = r-1
        doc_id = sheet.cell(row = r, column = 2).value    
        if doc_id in relevant_docs:

            summation += 1        
        temp_recall_dict[doc_id] = 1.0 * summation/relevant_docs_count[query_id]
        temp_precision_dict[doc_id] = 1.0 * summation/rank

    document_recall_dict[query_id] = temp_recall_dict
    document_precision_dict[query_id] = temp_precision_dict

#=============================================================================================
# computing average precision for each query.

MAP_query = 0
MAP_search_engine = 0

for query_id in query_ids:
    rel_docs_retrieved = 0
    MAP_query = 0
    filename = 'Query_' + query_id + '_results.xlsx'
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name('Sheet1')
    relevant_docs = query_relevance_levels.get(query_id).keys()
    for r in range(2, sheet.max_row + 1):    
        doc_id = sheet.cell(row = r, column = 2).value    
        if doc_id in relevant_docs:
            rel_docs_retrieved += 1
            MAP_query += document_precision_dict.get(query_id).get(doc_id)
    MAP_query = 1.0*MAP_query/len(relevant_docs)
    print("MAP - Query_id : " + query_id + " is : " + str(MAP_query))        
    MAP_search_engine += MAP_query     

# computing MAP for BM25 document retrieval algorithm

MAP_search_engine = MAP_search_engine/len(query_ids)
print(MAP_search_engine)

#=============================================================================================
# computing DCG for each document at specified ranks.

dcg_at_rank = dict((query_id, {}) for query_id in query_ids)

for query_id in query_ids:
    dcg_at_p = 0
    filename = 'Query_' + query_id + '_results.xlsx'
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name('Sheet1')
    relevant_docs = query_relevance_levels.get(query_id).keys()

    # computing dcg for rank-1 document    
    first_doc = sheet.cell(row = 2, column = 2).value    
    if first_doc in relevant_docs:        
        rel1 = int(query_relevance_levels.get(query_id).get(first_doc))
        dcg_at_p = rel1
        
    dcg_at_rank[query_id][first_doc] = dcg_at_p

    # computing DCG for docs having rank > 1
    for r in range(3, sheet.max_row + 1):    
        doc_id = sheet.cell(row = r, column = 2).value    
        if doc_id in relevant_docs:
            rel = int(query_relevance_levels.get(query_id).get(doc_id))
            dcg_at_p += 1.0*rel/math.log((r-1),2)            
        dcg_at_rank[query_id][doc_id] = dcg_at_p

#=============================================================================================
# constructing IDCG

idcg_at_rank = dict((query_id, {}) for query_id in query_ids)

for query_id in query_ids:
    idcg_at_p = 0
    relevant_docs = query_relevance_levels.get(query_id).keys()
    no_of_relevant_docs = len(relevant_docs)
    filename = 'Query_' + query_id + '_results.xlsx'
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name('Sheet1')
    first_doc = sheet.cell(row = 2, column = 2).value
    idcg_at_p = 1
    idcg_at_rank[query_id][first_doc] = idcg_at_p
    count = 2
    for r in range(3, sheet.max_row + 1):    
        doc_id = sheet.cell(row = r, column = 2).value
        if count <= no_of_relevant_docs:
            idcg_at_p += 1.0*1/math.log((r-1),2)
            idcg_at_rank[query_id][doc_id] = idcg_at_p
        else:
            idcg_at_rank[query_id][doc_id] = idcg_at_p
        count += 1
      
#=============================================================================================
# computing NDCG = DCG/IDCG

ndcg_at_rank = dict((query_id, {}) for query_id in query_ids)

for query_id in query_ids:    
    filename = 'Query_' + query_id + '_results.xlsx'
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name('Sheet1')
    for r in range(2, sheet.max_row + 1):    
        doc_id = sheet.cell(row = r, column = 2).value
        if idcg_at_rank[query_id][doc_id] == 0:
            ndcg_at_rank[query_id][doc_id] = 0
        else:
            ndcg_at_rank[query_id][doc_id] = 1.0*dcg_at_rank[query_id][doc_id]/idcg_at_rank[query_id][doc_id]

#=============================================================================================
# printing values to excel file

wb_output = openpyxl.Workbook()

for query_id in query_ids:
    sheet_count = 0
    sheet_name = 'query_id_' + query_id
    wb_output.create_sheet(index=0, title=sheet_name)
    sheet_output = wb_output.get_sheet_by_name(sheet_name)
    sheet_output['A1'] = 'RANK'
    sheet_output['B1'] = 'DOC_ID'
    sheet_output['C1'] = 'DOC_SCORE'
    sheet_output['D1'] = 'RELEVANCE_LEVEL'
    sheet_output['E1'] = 'PRECISION'
    sheet_output['F1'] = 'RECALL'
    sheet_output['G1'] = 'NDCG'
    ###
    #sheet_output['J1'] = 'DCG'
    #sheet_output['K1'] = 'IDCG'
    
    filename = 'Query_' + query_id + '_results.xlsx'
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name('Sheet1')
    
    relevant_docs = query_relevance_levels.get(query_id).keys()

    for r in range(2, sheet.max_row + 1):
        rank = r-1
        doc_id = sheet.cell(row = r, column = 2).value
        doc_score = sheet.cell(row = r, column = 3).value
        
        if doc_id in relevant_docs:
            relevance_level = query_relevance_levels[query_id][doc_id]
        else:
            relevance_level = 0
            
        precision = document_precision_dict[query_id][doc_id]
        recall = document_recall_dict[query_id][doc_id]
        ndcg = ndcg_at_rank[query_id][doc_id]
        
        sheet_output['A' + str(r)] = str(rank)
        sheet_output['B' + str(r)] = str(doc_id)
        sheet_output['C' + str(r)] = str(doc_score)
        sheet_output['D' + str(r)] = str(relevance_level)
        sheet_output['E' + str(r)] = str(precision)
        sheet_output['F' + str(r)] = str(recall)
        sheet_output['G' + str(r)] = str(ndcg)
        ##
        #sheet_output['J' + str(r)] = str(dcg_at_rank[query_id][doc_id])
        #sheet_output['K' + str(r)] = str(idcg_at_rank[query_id][doc_id])
        
            
    sheet_count += 1

wb_output.save('HW5_output.xlsx')

#=============================================================================================
# writing average precision values and MAP to text file

with open('MAP_and_P@K_values.txt', 'w') as f:
    f.write('MAP for Search Engine : ' + str(MAP_search_engine) + '\n')
    for query_id in query_ids:
        filename = 'Query_' + query_id + '_results.xlsx'
        wb = openpyxl.load_workbook(filename)
        sheet = wb.get_sheet_by_name('Sheet1')
        doc_20 = sheet.cell(row = 21, column = 2).value
        p_at_20 = document_precision_dict[query_id][doc_20]
        f.write('P@20 for Query_id - ' + query_id + ' : ' + str(p_at_20) + '\n')

f.close()        
        
#=============================================================================================        


























        


    






















































        
    
    


    
